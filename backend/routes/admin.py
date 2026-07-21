"""
Admin Routes — PIN-based authentication and dashboard data.

Security model:
  - Admin enters PIN "1234" to authenticate.
  - Server returns a random session token (stored in-memory).
  - All subsequent admin API calls must include the token in the Authorization header.
  - Tokens expire after 24 hours.

Note: This is a simple auth scheme suitable for internal tools.
For production, consider JWT tokens or OAuth.
"""

import io
import secrets
import time
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Header
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from database import get_db, Promoter, Submission, ValidUsername
from worker import remove_from_cache
from models import (
    AdminLoginRequest,
    AdminLoginResponse,
    AdminStatsResponse,
    AdminSubmission,
    BatchDeleteRequest,
)
from config import ADMIN_PIN, ADMIN_TOKEN_EXPIRY

router = APIRouter()

# ──────────────────────────────────────────────
# In-memory token store (maps token → creation timestamp)
# For production, use Redis or a database table.
# ──────────────────────────────────────────────
_admin_tokens: dict[str, float] = {}


def _generate_admin_token() -> str:
    """Generate a cryptographically secure session token."""
    token = secrets.token_hex(32)
    _admin_tokens[token] = time.time()
    return token


def verify_admin_token(authorization: Optional[str] = Header(None)) -> bool:
    """
    FastAPI dependency that verifies the admin token from the Authorization header.
    Expected format: "Bearer <token>"
    """
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization header required.")

    token = authorization.replace("Bearer ", "")

    if token not in _admin_tokens:
        raise HTTPException(status_code=401, detail="Invalid or expired token.")

    # Check token expiry
    created_at = _admin_tokens[token]
    if time.time() - created_at > ADMIN_TOKEN_EXPIRY:
        del _admin_tokens[token]
        raise HTTPException(status_code=401, detail="Token expired. Please log in again.")

    return True


# ──────────────────────────────────────────────
# Endpoints
# ──────────────────────────────────────────────

@router.post("/admin/login", response_model=AdminLoginResponse)
async def admin_login(request: AdminLoginRequest):
    """
    Authenticate admin with PIN code.
    Returns a session token on success.
    """
    if request.pin == ADMIN_PIN:
        token = _generate_admin_token()
        return AdminLoginResponse(
            success=True,
            token=token,
            message="Login successful.",
        )

    return AdminLoginResponse(
        success=False,
        token=None,
        message="Invalid PIN. Please try again.",
    )


# SQLite expression for the MYT calendar day of a submission (UTC + 8h)
_MYT_DAY = func.strftime("%Y-%m-%d", func.datetime(Submission.created_at, "+8 hours"))


@router.get("/admin/stats", response_model=AdminStatsResponse)
async def get_admin_stats(
    status_filter: Optional[str] = None,
    promoter_filter: Optional[str] = None,
    event_filter: Optional[str] = None,
    day_filter: Optional[str] = None,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin_token),
):
    """
    Get admin dashboard statistics and filtered submission list.

    Query params:
      - status_filter: Filter by submission status (valid|duplicate|ocr_failed)
      - promoter_filter: Search by promoter name (partial match)
      - event_filter: Filter by event/activation (exact match)
      - day_filter: Filter by MYT calendar day ("YYYY-MM-DD")
    """
    # Event + day narrow the whole scope (counts AND list); status/promoter
    # further narrow only the list so the stat cards show per-status totals
    # within the selected event/day.
    def scoped(q):
        if event_filter:
            q = q.filter(Submission.event == event_filter)
        if day_filter:
            q = q.filter(_MYT_DAY == day_filter)
        return q

    total_promoters = db.query(func.count(Promoter.id)).scalar() or 0
    total_submissions = scoped(db.query(func.count(Submission.id))).scalar() or 0
    total_valid = scoped(db.query(func.count(Submission.id)).filter(Submission.status == "valid")).scalar() or 0
    total_duplicate = scoped(db.query(func.count(Submission.id)).filter(Submission.status == "duplicate")).scalar() or 0
    total_ocr_failed = scoped(db.query(func.count(Submission.id)).filter(Submission.status == "ocr_failed")).scalar() or 0

    # ── Build filtered submission list ──
    query = scoped(
        db.query(Submission, Promoter)
        .join(Promoter, Promoter.id == Submission.promoter_id)
    ).order_by(Submission.created_at.desc())

    if status_filter and status_filter in ("valid", "duplicate", "ocr_failed"):
        query = query.filter(Submission.status == status_filter)
    if promoter_filter:
        query = query.filter(Promoter.name.ilike(f"%{promoter_filter}%"))

    submissions_data = query.limit(200).all()

    submissions = []
    for sub, promoter in submissions_data:
        submissions.append(
            AdminSubmission(
                id=sub.id,
                promoter_name=promoter.name,
                ic_number=promoter.ic_number,
                extracted_username=sub.extracted_username,
                full_name=sub.full_name,
                member_id=sub.member_id,
                event=sub.event,
                status=sub.status,
                image_path=sub.image_path,
                created_at=sub.created_at.isoformat() if sub.created_at else "",
                ocr_time=sub.ocr_time,
                rule_time=sub.rule_time,
                matching_time=sub.matching_time,
                total_time=sub.total_time,
                ocr_confidence=sub.ocr_confidence,
                candidate_score=sub.candidate_score,
                matched_name=sub.matched_name,
                similarity=sub.similarity,
                llm_used=bool(sub.llm_used) if sub.llm_used is not None else False,
            )
        )

    # Distinct events/days across the FULL dataset (so filters stay switchable)
    events = [r[0] for r in db.query(Submission.event).filter(Submission.event.isnot(None)).distinct().all() if r[0]]
    days = [r[0] for r in db.query(_MYT_DAY).distinct().all() if r[0]]
    events.sort()
    days.sort(reverse=True)

    return AdminStatsResponse(
        total_promoters=total_promoters,
        total_submissions=total_submissions,
        total_valid=total_valid,
        total_duplicate=total_duplicate,
        total_ocr_failed=total_ocr_failed,
        submissions=submissions,
        events=events,
        days=days,
    )


def _safe(value):
    """Neutralize spreadsheet formula injection. A value beginning with a
    formula trigger (= + - @) or a control char could execute when the file is
    opened (or, via openpyxl, be written as a real formula). Prefix a quote so
    it renders strictly as text. Legitimate names/IDs never start with these."""
    if isinstance(value, str) and value and value[0] in ("=", "+", "-", "@", "\t", "\r", "\n"):
        return "'" + value
    return value


def _fmt_myt(dt) -> str:
    """Format a stored (UTC) datetime as Malaysia local time for the report."""
    if not dt:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    myt = dt.astimezone(timezone(timedelta(hours=8)))
    return myt.strftime("%Y-%m-%d %H:%M")


@router.get("/admin/export")
async def export_excel(
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin_token),
):
    """
    One-click Excel export of all campaign data.
    Three sheets: Signups (valid only), All Submissions, Promoters.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0066CC")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A2"

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 45)

    # Pull all submissions joined with promoter, newest first
    rows = (
        db.query(Submission, Promoter)
        .join(Promoter, Promoter.id == Submission.promoter_id)
        .order_by(Submission.created_at.desc())
        .all()
    )

    def myt_day(dt) -> str:
        if not dt:
            return ""
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone(timedelta(hours=8))).strftime("%Y-%m-%d")

    # ── Sheet 1: Summary — valid signups by Event × Day (payout view) ──
    ws0 = wb.active
    ws0.title = "Summary"
    ws0.append(["Event", "Day (MYT)", "Valid Signups", "Total Uploads"])
    agg: dict = {}
    for sub, _p in rows:
        key = (sub.event or "(no event)", myt_day(sub.created_at))
        cell = agg.setdefault(key, {"valid": 0, "total": 0})
        cell["total"] += 1
        if sub.status == "valid":
            cell["valid"] += 1
    for (ev, day) in sorted(agg.keys()):
        ws0.append([_safe(ev), day, agg[(ev, day)]["valid"], agg[(ev, day)]["total"]])
    style_header(ws0, 4)
    autosize(ws0)

    # ── Sheet 2: Signups (valid only) — the payout-relevant rows ──
    ws1 = wb.create_sheet("Signups")
    ws1.append(["No.", "Event", "Day (MYT)", "Promoter", "Full Name", "Member ID", "Time (MYT)"])
    n = 0
    for sub, promoter in rows:
        if sub.status != "valid":
            continue
        n += 1
        ws1.append([
            n,
            _safe(sub.event or ""),
            myt_day(sub.created_at),
            _safe(promoter.name),
            _safe(sub.full_name or sub.extracted_username or ""),
            _safe(sub.member_id or ""),
            _fmt_myt(sub.created_at),
        ])
    style_header(ws1, 7)
    autosize(ws1)

    # ── Sheet 3: All Submissions ──
    ws2 = wb.create_sheet("All Submissions")
    ws2.append([
        "ID", "Event", "Day (MYT)", "Promoter", "IC Number", "Full Name", "Member ID",
        "Status", "Matched With", "Similarity %", "OCR Confidence", "Date/Time (MYT)",
    ])
    for sub, promoter in rows:
        ws2.append([
            sub.id,
            _safe(sub.event or ""),
            myt_day(sub.created_at),
            _safe(promoter.name),
            _safe(promoter.ic_number),
            _safe(sub.full_name or sub.extracted_username or ""),
            _safe(sub.member_id or ""),
            sub.status,
            _safe(sub.matched_name or ""),
            round(sub.similarity, 1) if sub.similarity is not None else "",
            round(sub.ocr_confidence, 2) if sub.ocr_confidence is not None else "",
            _fmt_myt(sub.created_at),
        ])
    style_header(ws2, 12)
    autosize(ws2)

    # ── Sheet 4: Promoters (with valid signup counts) ──
    ws3 = wb.create_sheet("Promoters")
    ws3.append(["Promoter", "IC Number", "Gender", "Valid Signups", "Joined (MYT)"])
    counts = dict(
        db.query(ValidUsername.promoter_id, func.count(ValidUsername.id))
        .group_by(ValidUsername.promoter_id)
        .all()
    )
    promoters = db.query(Promoter).order_by(Promoter.name).all()
    for p in promoters:
        ws3.append([
            _safe(p.name),
            _safe(p.ic_number),
            _safe(p.gender or ""),
            counts.get(p.id, 0),
            _fmt_myt(p.created_at),
        ])
    style_header(ws3, 5)
    autosize(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    stamp = datetime.now(timezone(timedelta(hours=8))).strftime("%Y%m%d_%H%M")
    filename = f"promoter-data-{stamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/admin/submission/{submission_id}")
async def delete_submission(
    submission_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin_token),
):
    """
    Delete a submission by ID.
    If the submission was 'valid', this also deletes the username from valid_usernames
    to release the unique constraint.
    Also deletes the physical image file on disk.
    """
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    # 1. Delete from valid_usernames if it exists, and clear from in-memory cache
    valid_entry = db.query(ValidUsername).filter(ValidUsername.submission_id == submission_id).first()
    if valid_entry:
        remove_from_cache(valid_entry.username)
        db.delete(valid_entry)

    # 2. Try deleting physical file on disk
    try:
        from config import UPLOAD_DIR
        file_path = UPLOAD_DIR / submission.image_path
        
        # Release any open Python file descriptors
        import gc
        gc.collect()
        
        if file_path.exists() and file_path.is_file():
            try:
                file_path.unlink()
            except PermissionError:
                # Fallback retry for Windows file locks
                import time
                time.sleep(0.1)
                gc.collect()
                file_path.unlink()
    except Exception as e:
        print(f"[Admin] Failed to delete image file: {str(e)}")

    # 3. Delete submission record
    db.delete(submission)
    db.commit()

    return {"success": True, "message": "Submission deleted successfully."}


@router.post("/admin/submissions/batch-delete")
async def delete_submissions_batch(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin_token),
):
    """
    Delete multiple submissions by ID.
    Releases username constraints from valid_usernames and deletes files on disk.
    """
    deleted_count = 0
    errors = []
    
    # Fetch all submissions to delete
    submissions = db.query(Submission).filter(Submission.id.in_(request.ids)).all()
    
    for sub in submissions:
        # Delete from valid_usernames and clear from in-memory cache
        valid_entry = db.query(ValidUsername).filter(ValidUsername.submission_id == sub.id).first()
        if valid_entry:
            remove_from_cache(valid_entry.username)
            db.delete(valid_entry)
        
        # Delete file on disk
        try:
            from config import UPLOAD_DIR
            file_path = UPLOAD_DIR / sub.image_path
            
            import gc
            gc.collect()
            
            if file_path.exists() and file_path.is_file():
                try:
                    file_path.unlink()
                except PermissionError:
                    import time
                    time.sleep(0.1)
                    gc.collect()
                    file_path.unlink()
        except Exception as e:
            err_msg = f"Failed to delete file for submission {sub.id}: {str(e)}"
            print(f"[Admin] {err_msg}")
            errors.append(err_msg)
            
        # Delete record
        db.delete(sub)
        deleted_count += 1
        
    db.commit()
    return {
        "success": True, 
        "message": f"Successfully deleted {deleted_count} submissions.",
        "errors": errors
    }


@router.get("/admin/promoters")
async def get_admin_promoters(
    db: Session = Depends(get_db),
    _: bool = Depends(verify_admin_token),
):
    """
    Retrieve all promoters, including their name, IC number, and gender.
    """
    promoters = db.query(Promoter).order_by(Promoter.name.asc()).all()
    return [
        {
            "id": p.id,
            "name": p.name,
            "ic_number": p.ic_number,
            "gender": p.gender or "unknown",
            "avatar": p.avatar,
            "created_at": p.created_at.isoformat() if p.created_at else "",
        }
        for p in promoters
    ]

